"""OMR scanning workflow tests.

Covers the eligibility resolver (Decision D2 in docs/omr-implementation.md) and
answer-key management.

Exam subjects are built directly through the ORM rather than over HTTP because
the scenarios here need mark-scheme shapes (multi-section, MCQ-only, no
sections) that would otherwise take several setup calls each. The session is
committed so the app's own request-scoped sessions can see the rows.
"""

import io
import re
import uuid
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace

import pytest
from httpx import ASGITransport, AsyncClient
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import NotFoundException, ValidationException
from app.main import app
from app.modules.academic.models import AcademicYear, Class, Section, StudentEnrollment, Subject
from app.modules.auth.models import Role, User
from app.modules.exams.models import Exam, ExamSubject, ExamSubjectSection
from app.modules.omr.service import resolve_mcq_marks
from app.modules.students.models import Student
from app.modules.teachers.models import Teacher

DEMO_TEACHER_EMAIL = "teacher@codexedumine.test"

FIXTURES = Path(__file__).parent / "fixtures" / "omr"

# Values the CV engine reads off the committed fixture sheets (pinned by
# tests/test_omr_engine.py). The roll numbers are zero-padded to the sheet's six
# columns; enrollment rolls are stored unpadded, which is exactly the
# normalization tier D3 describes.
SHEET_FACTS = {
    "omr1": {"class": 9, "roll": "011223", "set": "Ga"},
    "omr2": {"class": 10, "roll": "028637", "set": "Nga"},
    "omr3": {"class": 10, "roll": "374789", "set": "Gha"},
}
ALL_SET_CODES = ["Ka", "Kha", "Ga", "Gha", "Nga", "Cha"]


async def _demo_teacher(db: AsyncSession) -> Teacher:
    """The teacher behind role_clients["teacher"] — ownership tests depend on it
    being that exact teacher, not merely any teacher row."""
    teacher = (
        await db.execute(
            select(Teacher).join(User, User.id == Teacher.user_id).where(User.email == DEMO_TEACHER_EMAIL)
        )
    ).scalars().first()
    assert teacher is not None, f"demo seed should provide {DEMO_TEACHER_EMAIL}"
    return teacher


async def _make_exam_subject(
    db: AsyncSession,
    suffix: str,
    *,
    full_marks: int = 100,
    sections: list[tuple[str, int]] | None = None,
    teacher: Teacher | None = None,
    class_order: int = 9,
    subject_code: str | None = None,
) -> ExamSubject:
    """Build the minimal exam-subject chain the resolver needs.

    `sections` is a list of (name, full_marks); omit it for a subject with no
    mark-scheme breakdown.
    """
    teacher = teacher or await _demo_teacher(db)

    year = AcademicYear(
        name=f"Y{suffix}"[:20],
        start_date=date(2026, 1, 1),
        end_date=date(2026, 12, 31),
        is_active=False,  # a partial unique index allows only one active year
    )
    class_entity = Class(name=f"Class {suffix}", numeric_order=class_order)
    subject = Subject(name=f"Subject {suffix}", code=(subject_code or f"S{suffix}")[:20])
    db.add_all([year, class_entity, subject])
    await db.flush()

    exam = Exam(
        academic_year_id=year.id,
        name=f"Exam {suffix}",
        start_date=date(2026, 6, 1),
        end_date=date(2026, 6, 10),
        created_by=teacher.user_id,
    )
    db.add(exam)
    await db.flush()

    now = datetime.now(timezone.utc)
    exam_subject = ExamSubject(
        exam_id=exam.id,
        class_id=class_entity.id,
        subject_id=subject.id,
        teacher_id=teacher.id,
        full_marks=full_marks,
        pass_marks=int(full_marks * 0.33),
        question_deadline=now + timedelta(hours=2),
        marks_deadline=now + timedelta(hours=3),
    )
    db.add(exam_subject)
    await db.flush()

    for order, (name, section_marks) in enumerate(sections or []):
        db.add(
            ExamSubjectSection(
                exam_subject_id=exam_subject.id,
                name=name,
                full_marks=section_marks,
                pass_marks=int(section_marks * 0.33),
                display_order=order,
            )
        )
    await db.commit()
    return exam_subject


def _simple_key(total: int = 4) -> dict:
    options = ["Ka", "Kha", "Ga", "Gha"]
    return {
        "total_questions": total,
        "answers": {str(i): options[(i - 1) % 4] for i in range(1, total + 1)},
    }


# ==============================================================================
# Eligibility resolver (Decision D2)
# ==============================================================================


async def test_subject_without_sections_uses_flat_full_marks(db_session, unique_suffix):
    exam_subject = await _make_exam_subject(db_session, unique_suffix, full_marks=40)

    resolution = await resolve_mcq_marks(db_session, exam_subject.id)

    assert resolution.mcq_full_marks == 40
    assert resolution.source == "whole_subject"
    assert resolution.section_id is None


async def test_single_mcq_section_supplies_the_ceiling(db_session, unique_suffix):
    exam_subject = await _make_exam_subject(
        db_session, unique_suffix, full_marks=50, sections=[("MCQ", 50)]
    )

    resolution = await resolve_mcq_marks(db_session, exam_subject.id)

    assert resolution.mcq_full_marks == 50
    assert resolution.source == "section"
    assert resolution.section_name == "MCQ"
    assert resolution.section_id is not None


@pytest.mark.parametrize("name", ["mcq", "Mcq", "  MCQ  "])
async def test_mcq_section_name_matches_case_insensitively(db_session, unique_suffix, name):
    exam_subject = await _make_exam_subject(
        db_session, f"{unique_suffix}{abs(hash(name)) % 997}", full_marks=30, sections=[(name, 30)]
    )

    resolution = await resolve_mcq_marks(db_session, exam_subject.id)

    assert resolution.mcq_full_marks == 30
    assert resolution.source == "section"


async def test_multi_section_subject_is_rejected(db_session, unique_suffix):
    """D2 option A: an MCQ subtotal has nowhere to live in the flat
    exam_results.marks_obtained, so mixed subjects are refused outright."""
    exam_subject = await _make_exam_subject(
        db_session, unique_suffix, full_marks=100, sections=[("MCQ", 40), ("CQ", 60)]
    )

    with pytest.raises(ValidationException) as exc:
        await resolve_mcq_marks(db_session, exam_subject.id)

    assert exc.value.message == "OMR scanning is only supported for MCQ-only exam subjects"
    assert "CQ" in exc.value.details[0]["issue"]


async def test_subject_with_sections_but_no_mcq_is_rejected(db_session, unique_suffix):
    exam_subject = await _make_exam_subject(
        db_session, unique_suffix, full_marks=100, sections=[("Written", 70), ("Practical", 30)]
    )

    with pytest.raises(ValidationException) as exc:
        await resolve_mcq_marks(db_session, exam_subject.id)

    message = exc.value.message
    assert "no MCQ section" in message
    # The message must name what the subject actually has, so the fix is obvious.
    assert "Written" in message and "Practical" in message


async def test_unknown_exam_subject_raises_not_found(db_session):
    with pytest.raises(NotFoundException):
        await resolve_mcq_marks(db_session, uuid.uuid4())


# ==============================================================================
# Eligibility endpoint
# ==============================================================================


async def test_eligibility_endpoint_reports_eligible_subject(role_clients, db_session, unique_suffix):
    exam_subject = await _make_exam_subject(
        db_session, unique_suffix, full_marks=40, sections=[("MCQ", 40)]
    )

    resp = await role_clients["teacher"].get(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/eligibility"
    )

    assert resp.status_code == 200, resp.text
    data = resp.json()["data"]
    assert data["eligible"] is True
    assert data["mcq_full_marks"] == 40
    assert data["source"] == "section"
    assert data["answer_key_set_codes"] == []


async def test_eligibility_endpoint_reports_reason_instead_of_erroring(
    role_clients, db_session, unique_suffix
):
    """The UI must be able to list ineligible subjects with their reason, so D2's
    rejection comes back as data here rather than as a 422."""
    exam_subject = await _make_exam_subject(
        db_session, unique_suffix, full_marks=100, sections=[("MCQ", 40), ("CQ", 60)]
    )

    resp = await role_clients["teacher"].get(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/eligibility"
    )

    assert resp.status_code == 200, resp.text
    data = resp.json()["data"]
    assert data["eligible"] is False
    assert data["reason"] == "OMR scanning is only supported for MCQ-only exam subjects"
    assert data["mcq_full_marks"] is None


# ==============================================================================
# Answer keys
# ==============================================================================


async def test_simple_answer_key_round_trips(role_clients, db_session, unique_suffix):
    exam_subject = await _make_exam_subject(db_session, unique_suffix, full_marks=40)
    teacher = role_clients["teacher"]

    resp = await teacher.put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Ka", json=_simple_key(4)
    )
    assert resp.status_code == 201, resp.text
    assert resp.json()["data"]["answers"] == {"1": "Ka", "2": "Kha", "3": "Ga", "4": "Gha"}

    resp = await teacher.get(f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys")
    assert resp.status_code == 200, resp.text
    keys = resp.json()["data"]
    assert len(keys) == 1
    assert keys[0]["set_code"] == "Ka"
    assert keys[0]["total_questions"] == 4


async def test_extended_answer_key_round_trips(role_clients, db_session, unique_suffix):
    exam_subject = await _make_exam_subject(db_session, unique_suffix, full_marks=40)

    resp = await role_clients["teacher"].put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Kha",
        json={
            "total_questions": 2,
            "answers": {
                "1": {"correct": "ka", "marks": 2, "negative": 0.5},
                "2": {"correct": "GHA", "marks": 1, "negative": 0.25},
            },
        },
    )

    assert resp.status_code == 201, resp.text
    answers = resp.json()["data"]["answers"]
    assert answers["1"] == {"correct": "Ka", "marks": 2.0, "negative": 0.5}
    assert answers["2"] == {"correct": "Gha", "marks": 1.0, "negative": 0.25}


async def test_option_casing_is_normalized(role_clients, db_session, unique_suffix):
    exam_subject = await _make_exam_subject(db_session, unique_suffix, full_marks=40)

    resp = await role_clients["teacher"].put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Ka",
        json={"total_questions": 4, "answers": {"1": "ka", "2": "KHA", "3": "gA", "4": " gha "}},
    )

    assert resp.status_code == 201, resp.text
    assert resp.json()["data"]["answers"] == {"1": "Ka", "2": "Kha", "3": "Ga", "4": "Gha"}


async def test_two_set_codes_coexist_on_one_subject(role_clients, db_session, unique_suffix):
    exam_subject = await _make_exam_subject(db_session, unique_suffix, full_marks=40)
    teacher = role_clients["teacher"]

    for set_code in ("Ka", "Nga"):
        resp = await teacher.put(
            f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/{set_code}", json=_simple_key(4)
        )
        assert resp.status_code == 201, resp.text

    resp = await teacher.get(f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys")
    assert sorted(k["set_code"] for k in resp.json()["data"]) == ["Ka", "Nga"]


async def test_reputting_a_set_code_replaces_rather_than_duplicating(
    role_clients, db_session, unique_suffix
):
    """PUT is create-or-replace per §3.2, so a second PUT to the same set code
    updates the existing key — a teacher fixing a typo must not have to delete
    the key first. The (exam_subject_id, set_code) unique constraint guarantees
    only ever one row per set code."""
    exam_subject = await _make_exam_subject(db_session, unique_suffix, full_marks=40)
    teacher = role_clients["teacher"]

    resp = await teacher.put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Ka", json=_simple_key(4)
    )
    assert resp.status_code == 201
    original_id = resp.json()["data"]["id"]

    resp = await teacher.put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Ka",
        json={"total_questions": 2, "answers": {"1": "Gha", "2": "Ga"}},
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["message"] == "Answer key replaced"
    assert resp.json()["data"]["id"] == original_id

    resp = await teacher.get(f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys")
    keys = resp.json()["data"]
    assert len(keys) == 1
    assert keys[0]["total_questions"] == 2
    assert keys[0]["answers"] == {"1": "Gha", "2": "Ga"}


async def test_answer_key_can_be_deleted(role_clients, db_session, unique_suffix):
    exam_subject = await _make_exam_subject(db_session, unique_suffix, full_marks=40)
    teacher = role_clients["teacher"]

    resp = await teacher.put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Ka", json=_simple_key(4)
    )
    key_id = resp.json()["data"]["id"]

    assert (await teacher.delete(f"/api/v1/omr/answer-keys/{key_id}")).status_code == 200
    resp = await teacher.get(f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys")
    assert resp.json()["data"] == []


# --- Validation ---------------------------------------------------------------


@pytest.mark.parametrize(
    "payload,expected_fragment",
    [
        ({"total_questions": 4, "answers": {"1": "Ka", "2": "Kha", "4": "Gha"}}, "missing 3"),
        ({"total_questions": 2, "answers": {"1": "Ka", "2": "Kha", "5": "Ga"}}, "unexpected 5"),
        ({"total_questions": 3, "answers": {"1": "Ka", "2": "Kha"}}, "missing 3"),
        ({"total_questions": 2, "answers": {"1": "Ka", "x": "Kha"}}, "not a number"),
    ],
)
async def test_non_contiguous_question_numbers_are_rejected(
    role_clients, db_session, unique_suffix, payload, expected_fragment
):
    exam_subject = await _make_exam_subject(
        db_session, f"{unique_suffix}{abs(hash(expected_fragment)) % 997}", full_marks=40
    )

    resp = await role_clients["teacher"].put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Ka", json=payload
    )

    assert resp.status_code == 422, resp.text
    assert expected_fragment in resp.text


@pytest.mark.parametrize("bad_option", ["A", "Kaa", "", "1"])
async def test_invalid_option_is_rejected(role_clients, db_session, unique_suffix, bad_option):
    exam_subject = await _make_exam_subject(
        db_session, f"{unique_suffix}{abs(hash(bad_option)) % 997}", full_marks=40
    )

    resp = await role_clients["teacher"].put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Ka",
        json={"total_questions": 1, "answers": {"1": bad_option}},
    )

    assert resp.status_code == 422, resp.text
    assert "not a valid option" in resp.text


async def test_invalid_set_code_is_rejected(role_clients, db_session, unique_suffix):
    exam_subject = await _make_exam_subject(db_session, unique_suffix, full_marks=40)

    resp = await role_clients["teacher"].put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Zed", json=_simple_key(4)
    )

    assert resp.status_code == 422, resp.text
    assert "not a valid set code" in resp.text


async def test_total_questions_is_bounded(role_clients, db_session, unique_suffix):
    exam_subject = await _make_exam_subject(db_session, unique_suffix, full_marks=40)

    resp = await role_clients["teacher"].put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Ka",
        json={"total_questions": 201, "answers": {"1": "Ka"}},
    )

    assert resp.status_code == 422, resp.text


async def test_answer_key_rejected_for_multi_section_subject(role_clients, db_session, unique_suffix):
    exam_subject = await _make_exam_subject(
        db_session, unique_suffix, full_marks=100, sections=[("MCQ", 40), ("CQ", 60)]
    )

    resp = await role_clients["teacher"].put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Ka", json=_simple_key(4)
    )

    assert resp.status_code == 422, resp.text
    assert resp.json()["message"] == "OMR scanning is only supported for MCQ-only exam subjects"


# --- Access control -----------------------------------------------------------


async def test_admin_and_principal_bypass_ownership(role_clients, db_session, unique_suffix):
    exam_subject = await _make_exam_subject(db_session, unique_suffix, full_marks=40)

    resp = await role_clients["admin"].put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Ka", json=_simple_key(4)
    )
    assert resp.status_code == 201, resp.text

    resp = await role_clients["principal"].put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Kha", json=_simple_key(4)
    )
    assert resp.status_code == 201, resp.text


async def test_another_teacher_is_denied(role_clients, db_session, unique_suffix):
    """Ownership must be tested with a second teacher, not Admin — Admin and
    Principal deliberately bypass the check."""
    exam_subject = await _make_exam_subject(db_session, unique_suffix, full_marks=40)

    resp = await role_clients["admin"].post(
        "/api/v1/teachers",
        json={
            "full_name": "OMR Other Teacher",
            "email": f"omr.other.{unique_suffix}@codexedumine.test",
            "phone": f"01933{unique_suffix[:6]}",
            "date_of_birth": "1990-01-01",
            "joining_date": str(date.today()),
        },
    )
    assert resp.status_code == 201, resp.text
    created = resp.json()["data"]

    other = AsyncClient(transport=ASGITransport(app=app), base_url="http://test")
    try:
        resp = await other.post(
            "/api/v1/auth/login",
            json={"identifier": created["email"], "password": created["temporary_password"]},
        )
        assert resp.status_code == 200, resp.text

        resp = await other.put(
            f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Ka", json=_simple_key(4)
        )
        assert resp.status_code == 403, resp.text

        resp = await other.get(f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys")
        assert resp.status_code == 403, resp.text
    finally:
        await other.aclose()


@pytest.mark.parametrize("role", ["student", "guardian", "accountant", "receptionist"])
async def test_roles_without_omr_permissions_are_denied(
    role_clients, db_session, unique_suffix, role
):
    exam_subject = await _make_exam_subject(
        db_session, f"{unique_suffix}{abs(hash(role)) % 997}", full_marks=40
    )

    resp = await role_clients[role].get(f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys")

    assert resp.status_code == 403, resp.text


# --- Applied-batch lock -------------------------------------------------------


async def test_replace_and_delete_rejected_once_a_batch_is_applied(
    role_clients, db_session, unique_suffix
):
    """Once marks derived from a key are in the roster, the key is frozen — a
    stored mark must stay reproducible from the key that produced it."""
    from app.common.enums import OmrBatchStatus
    from app.modules.omr.models import OmrBatch

    exam_subject = await _make_exam_subject(db_session, unique_suffix, full_marks=40)
    teacher = role_clients["teacher"]

    resp = await teacher.put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Ka", json=_simple_key(4)
    )
    key_id = resp.json()["data"]["id"]

    demo_teacher = await _demo_teacher(db_session)
    db_session.add(
        OmrBatch(
            exam_subject_id=exam_subject.id,
            name="Applied batch",
            status=OmrBatchStatus.applied,
            template_name="plus_coaching_template",
            mcq_full_marks=40,
            uploaded_by=demo_teacher.user_id,
        )
    )
    await db_session.commit()

    resp = await teacher.put(
        f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/Ka", json=_simple_key(4)
    )
    assert resp.status_code == 409, resp.text
    assert "already has an OMR batch applied" in resp.json()["message"]

    resp = await teacher.delete(f"/api/v1/omr/answer-keys/{key_id}")
    assert resp.status_code == 409, resp.text


# ==============================================================================
# Batch upload, processing and matching (Phase 5)
# ==============================================================================


@pytest.fixture
def cloudinary_stub(monkeypatch):
    """Exercise the real cloudinary code path with the SDK's network calls stubbed."""
    from app.core import storage

    monkeypatch.setattr(storage.settings, "storage_provider", "cloudinary")
    monkeypatch.setattr(storage.settings, "cloudinary_cloud_name", "test-cloud")
    monkeypatch.setattr(storage.settings, "cloudinary_api_key", "test-key")
    monkeypatch.setattr(storage.settings, "cloudinary_api_secret", "test-secret")
    monkeypatch.setattr(storage.settings, "cloudinary_folder", "codex-edumine")

    uploaded: dict[str, bytes] = {}
    destroyed: list[str] = []

    def fake_upload(data, **kwargs):
        public_id = "/".join(p for p in (kwargs.get("folder", ""), kwargs["public_id"]) if p)
        uploaded[public_id] = data
        return {
            "secure_url": f"https://res.cloudinary.com/test-cloud/image/upload/v1/{public_id}",
            "public_id": public_id,
        }

    def fake_destroy(public_id, **kwargs):
        destroyed.append(public_id)
        return {"result": "ok"}

    def fake_fetch(url):
        """Serve back whatever fake_upload stored, keyed by the URL's public id."""
        public_id = url.split("/image/upload/v1/", 1)[1]
        if public_id not in uploaded:
            raise FileNotFoundError(url)
        return uploaded[public_id]

    monkeypatch.setattr("cloudinary.config", lambda **kwargs: None)
    monkeypatch.setattr("cloudinary.uploader.upload", fake_upload)
    monkeypatch.setattr("cloudinary.uploader.destroy", fake_destroy)
    monkeypatch.setattr(storage, "_fetch_remote_sync", fake_fetch)

    return SimpleNamespace(uploaded=uploaded, destroyed=destroyed)


async def _enroll_student(
    db: AsyncSession, *, suffix: str, section: Section, academic_year_id: uuid.UUID, roll: str
) -> Student:
    student_role = (await db.execute(select(Role).where(Role.name == "student"))).scalars().one()
    user = User(
        role_id=student_role.id,
        full_name=f"OMR Student {roll}",
        email=f"omr.student.{suffix}.{roll}@codexedumine.test",
        phone=f"019{uuid.uuid4().int % 100000000:08d}",
        password_hash="x" * 20,  # never logged in during these tests
    )
    db.add(user)
    await db.flush()

    student = Student(
        user_id=user.id,
        admission_number=f"ADM{suffix}{roll}"[:30],
        admission_date=date(2026, 1, 1),
    )
    db.add(student)
    await db.flush()

    db.add(
        StudentEnrollment(
            student_id=student.id,
            academic_year_id=academic_year_id,
            section_id=section.id,
            roll_number=roll,
        )
    )
    await db.flush()
    return student


async def _make_scannable_batch(
    db: AsyncSession,
    role_clients,
    suffix: str,
    *,
    class_order: int = 10,
    enrolled_rolls: tuple[str, ...] = (),
    full_marks: int = 40,
) -> SimpleNamespace:
    """An MCQ-only exam subject with answer keys for every set code, enrolled
    students, and a batch ready to receive sheets."""
    exam_subject = await _make_exam_subject(
        db, suffix, full_marks=full_marks, class_order=class_order, subject_code=f"SUBJ{suffix}"[:20]
    )
    exam = (await db.execute(select(Exam).where(Exam.id == exam_subject.exam_id))).scalars().one()

    section = Section(
        academic_year_id=exam.academic_year_id, class_id=exam_subject.class_id, name="A"
    )
    db.add(section)
    await db.flush()

    students = {
        roll: await _enroll_student(
            db, suffix=suffix, section=section, academic_year_id=exam.academic_year_id, roll=roll
        )
        for roll in enrolled_rolls
    }
    await db.commit()

    teacher = role_clients["teacher"]
    options = ["Ka", "Kha", "Ga", "Gha"]
    for set_code in ALL_SET_CODES:
        resp = await teacher.put(
            f"/api/v1/omr/exam-subjects/{exam_subject.id}/answer-keys/{set_code}",
            json={
                "total_questions": 40,
                "answers": {str(i): options[(i - 1) % 4] for i in range(1, 41)},
            },
        )
        assert resp.status_code == 201, resp.text

    resp = await teacher.post(
        "/api/v1/omr/batches",
        json={"exam_subject_id": str(exam_subject.id), "name": f"Batch {suffix}"},
    )
    assert resp.status_code == 201, resp.text

    return SimpleNamespace(
        exam_subject=exam_subject,
        batch_id=resp.json()["data"]["id"],
        section=section,
        students=students,
        client=teacher,
    )


def _sheet_files(*names: str) -> list[tuple[str, tuple[str, bytes, str]]]:
    return [
        ("images", (f"{name}.jpg", (FIXTURES / f"{name}.jpg").read_bytes(), "image/jpeg"))
        for name in names
    ]


# --- Batch creation -----------------------------------------------------------


async def test_batch_creation_snapshots_the_mcq_ceiling(role_clients, db_session, unique_suffix):
    setup = await _make_scannable_batch(db_session, role_clients, unique_suffix, full_marks=40)

    resp = await setup.client.get(f"/api/v1/omr/batches/{setup.batch_id}")

    assert resp.status_code == 200, resp.text
    data = resp.json()["data"]
    assert data["mcq_full_marks"] == 40
    assert data["status"] == "draft"
    assert data["template_name"] == "plus_coaching_template"
    assert data["sheet_count"] == 0


async def test_batch_creation_requires_an_answer_key(role_clients, db_session, unique_suffix):
    exam_subject = await _make_exam_subject(db_session, unique_suffix, full_marks=40)

    resp = await role_clients["teacher"].post(
        "/api/v1/omr/batches",
        json={"exam_subject_id": str(exam_subject.id), "name": "No key batch"},
    )

    assert resp.status_code == 422, resp.text
    assert "answer key" in resp.json()["message"].lower()


async def test_batch_creation_rejects_multi_section_subject(role_clients, db_session, unique_suffix):
    exam_subject = await _make_exam_subject(
        db_session, unique_suffix, full_marks=100, sections=[("MCQ", 40), ("CQ", 60)]
    )

    resp = await role_clients["teacher"].post(
        "/api/v1/omr/batches",
        json={"exam_subject_id": str(exam_subject.id), "name": "Mixed batch"},
    )

    assert resp.status_code == 422, resp.text
    assert resp.json()["message"] == "OMR scanning is only supported for MCQ-only exam subjects"


# --- Sheet upload -------------------------------------------------------------


async def test_uploading_the_fixture_sheets_scores_all_three(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session,
        role_clients,
        unique_suffix,
        class_order=10,
        enrolled_rolls=("11223", "28637", "374789"),
    )

    resp = await setup.client.post(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets", files=_sheet_files("omr1", "omr2", "omr3")
    )

    assert resp.status_code == 201, resp.text
    data = resp.json()["data"]
    assert data["rejected"] == []
    sheets = data["sheets"]
    assert len(sheets) == 3

    by_roll = {sheet["detected_roll"]: sheet for sheet in sheets}
    for name, facts in SHEET_FACTS.items():
        sheet = by_roll[facts["roll"]]
        assert sheet["detected_set_code"] == facts["set"], name
        assert sheet["detected_class"] == facts["class"], name
        assert sheet["alignment_method"] == "markers", name
        # Every sheet found its answer key and was scored.
        assert sheet["marks_obtained"] is not None, name
        assert 0 <= sheet["marks_obtained"] <= 40, name
        assert sheet["correct_count"] + sheet["wrong_count"] + sheet["blank_count"] + sheet[
            "multiple_count"
        ] == 40, name
        assert sheet["image_url"].startswith("https://res.cloudinary.com/"), name

    # Rolls are stored unpadded; the sheet reads them zero-padded. All three
    # still resolve to a student via D3's normalization tiers.
    assert {s["match_status"] for s in sheets} == {"matched"}

    batch = data["batch"]
    assert batch["status"] == "processing"
    assert batch["sheet_count"] == 3
    assert batch["matched_count"] == 3
    assert batch["failed_count"] == 0


async def test_annotated_overlay_is_stored_alongside_the_original(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )

    resp = await setup.client.post(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets", files=_sheet_files("omr1")
    )

    sheet = resp.json()["data"]["sheets"][0]
    assert sheet["annotated_image_url"].startswith("https://res.cloudinary.com/")
    assert sheet["annotated_image_url"].endswith("_annotated")
    # Both assets live under the batch's folder so a batch delete can find them.
    prefix = f"codex-edumine/omr/{setup.exam_subject.id}/{setup.batch_id}/"
    assert all(key.startswith(prefix) for key in cloudinary_stub.uploaded)
    assert len(cloudinary_stub.uploaded) == 2


async def test_unmatched_roll_is_flagged_for_review(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("99999",)
    )

    resp = await setup.client.post(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets", files=_sheet_files("omr1")
    )

    sheet = resp.json()["data"]["sheets"][0]
    assert sheet["match_status"] == "unmatched"
    assert sheet["status"] == "needs_review"
    assert sheet["student_id"] is None
    # It was still read and scored — only the identity is in question.
    assert sheet["marks_obtained"] is not None


async def test_two_sheets_for_one_student_yield_matched_then_duplicate(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )

    resp = await setup.client.post(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets", files=_sheet_files("omr1", "omr1")
    )

    assert resp.status_code == 201, resp.text
    statuses = [sheet["match_status"] for sheet in resp.json()["data"]["sheets"]]
    assert statuses == ["matched", "duplicate"]

    duplicate = resp.json()["data"]["sheets"][1]
    assert duplicate["status"] == "needs_review"
    # The duplicate still records who it looked like, so a reviewer can see the clash.
    assert duplicate["student_id"] == str(setup.students["11223"].id)


async def test_class_mismatch_flags_review_without_changing_the_match(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    """omr1 is marked class 9; this batch is for class 10. The sheet was probably
    scanned into the wrong batch — flag it, but keep the roll-based match."""
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=10, enrolled_rolls=("11223",)
    )

    resp = await setup.client.post(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets", files=_sheet_files("omr1")
    )

    sheet = resp.json()["data"]["sheets"][0]
    assert sheet["match_status"] == "matched"
    assert sheet["status"] == "needs_review"
    assert "class 9" in sheet["review_note"] and "class 10" in sheet["review_note"]


async def test_corrupt_image_fails_only_its_own_sheet(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )

    files = _sheet_files("omr1")
    files.append(("images", ("broken.jpg", b"this is not an image at all", "image/jpeg")))

    resp = await setup.client.post(f"/api/v1/omr/batches/{setup.batch_id}/sheets", files=files)

    assert resp.status_code == 201, resp.text
    sheets = resp.json()["data"]["sheets"]
    assert len(sheets) == 2

    good, bad = sheets
    assert good["status"] in ("processed", "needs_review")
    assert bad["status"] == "failed"
    assert bad["error_message"]
    # The unreadable image is still stored, so Phase 6 can reprocess it.
    assert bad["image_url"].startswith("https://res.cloudinary.com/")
    assert resp.json()["data"]["batch"]["failed_count"] == 1


async def test_unsupported_file_type_is_rejected_without_a_row(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(db_session, role_clients, unique_suffix)

    resp = await setup.client.post(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets",
        files=[("images", ("scan.pdf", b"%PDF-1.4 fake", "application/pdf"))],
    )

    assert resp.status_code == 201, resp.text
    data = resp.json()["data"]
    assert data["sheets"] == []
    assert len(data["rejected"]) == 1
    assert "Unsupported file type" in data["rejected"][0]["reason"]
    assert cloudinary_stub.uploaded == {}


async def test_too_many_sheets_in_one_request_is_rejected(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    from app.modules.omr import service as omr_service

    setup = await _make_scannable_batch(db_session, role_clients, unique_suffix)
    tiny = (FIXTURES / "blank.jpg").read_bytes()
    limit = omr_service.settings.omr_max_sheets_per_request

    resp = await setup.client.post(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets",
        files=[("images", (f"s{i}.jpg", tiny, "image/jpeg")) for i in range(limit + 1)],
    )

    assert resp.status_code == 422, resp.text
    assert "at most" in resp.json()["message"]


# --- Listing and detail -------------------------------------------------------


async def test_sheets_can_be_listed_and_filtered(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    await setup.client.post(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets", files=_sheet_files("omr1", "omr2")
    )

    resp = await setup.client.get(f"/api/v1/omr/batches/{setup.batch_id}/sheets")
    assert resp.status_code == 200, resp.text
    assert len(resp.json()["data"]) == 2

    resp = await setup.client.get(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets", params={"match_status": "matched"}
    )
    matched = resp.json()["data"]
    assert len(matched) == 1
    assert matched[0]["detected_roll"] == "011223"
    assert matched[0]["student_name"] == "OMR Student 11223"


async def test_sheet_detail_includes_the_per_question_breakdown(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    resp = await setup.client.post(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets", files=_sheet_files("omr1")
    )
    sheet_id = resp.json()["data"]["sheets"][0]["id"]

    resp = await setup.client.get(f"/api/v1/omr/sheets/{sheet_id}")

    assert resp.status_code == 200, resp.text
    data = resp.json()["data"]
    assert len(data["answers"]) == 40
    assert len(data["score_details"]) == 40
    first = data["score_details"]["1"]
    assert set(first) >= {"student", "correct", "status", "marks"}


async def test_batches_are_listed_and_filterable(role_clients, db_session, unique_suffix):
    setup = await _make_scannable_batch(db_session, role_clients, unique_suffix)

    resp = await setup.client.get(
        "/api/v1/omr/batches", params={"exam_subject_id": str(setup.exam_subject.id)}
    )
    assert resp.status_code == 200, resp.text
    assert [b["id"] for b in resp.json()["data"]] == [setup.batch_id]

    resp = await setup.client.get(
        "/api/v1/omr/batches",
        params={"exam_subject_id": str(setup.exam_subject.id), "status": "applied"},
    )
    assert resp.json()["data"] == []


# --- Deletion -----------------------------------------------------------------


async def test_deleting_a_batch_removes_its_cloudinary_assets(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    await setup.client.post(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets", files=_sheet_files("omr1")
    )
    stored = set(cloudinary_stub.uploaded)
    assert len(stored) == 2

    resp = await setup.client.delete(f"/api/v1/omr/batches/{setup.batch_id}")

    assert resp.status_code == 200, resp.text
    assert set(cloudinary_stub.destroyed) == stored
    assert (await setup.client.get(f"/api/v1/omr/batches/{setup.batch_id}")).status_code == 404


async def test_applied_batch_cannot_be_deleted_or_extended(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    from app.common.enums import OmrBatchStatus
    from app.modules.omr.models import OmrBatch

    setup = await _make_scannable_batch(db_session, role_clients, unique_suffix)
    batch = (
        await db_session.execute(select(OmrBatch).where(OmrBatch.id == uuid.UUID(setup.batch_id)))
    ).scalars().one()
    batch.status = OmrBatchStatus.applied
    await db_session.commit()

    resp = await setup.client.delete(f"/api/v1/omr/batches/{setup.batch_id}")
    assert resp.status_code == 409, resp.text

    resp = await setup.client.post(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets", files=_sheet_files("omr1")
    )
    assert resp.status_code == 409, resp.text


# --- Access control -----------------------------------------------------------


async def test_another_teacher_cannot_touch_a_batch(role_clients, db_session, unique_suffix):
    setup = await _make_scannable_batch(db_session, role_clients, unique_suffix)

    resp = await role_clients["admin"].post(
        "/api/v1/teachers",
        json={
            "full_name": "OMR Batch Outsider",
            "email": f"omr.outsider.{unique_suffix}@codexedumine.test",
            "phone": f"01944{unique_suffix[:6]}",
            "date_of_birth": "1990-01-01",
            "joining_date": str(date.today()),
        },
    )
    created = resp.json()["data"]

    other = AsyncClient(transport=ASGITransport(app=app), base_url="http://test")
    try:
        await other.post(
            "/api/v1/auth/login",
            json={"identifier": created["email"], "password": created["temporary_password"]},
        )
        assert (await other.get(f"/api/v1/omr/batches/{setup.batch_id}")).status_code == 403
        assert (await other.delete(f"/api/v1/omr/batches/{setup.batch_id}")).status_code == 403
        # A batch they cannot see must not appear in their own listing either.
        resp = await other.get("/api/v1/omr/batches")
        assert resp.status_code == 200, resp.text
        assert setup.batch_id not in [b["id"] for b in resp.json()["data"]]
    finally:
        await other.aclose()


# ==============================================================================
# Review and correction (Phase 6)
# ==============================================================================


async def _upload(setup, *names: str) -> list[dict]:
    resp = await setup.client.post(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets", files=_sheet_files(*names)
    )
    assert resp.status_code == 201, resp.text
    return resp.json()["data"]["sheets"]


async def test_manual_assignment_flips_unmatched_to_manual(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("99999",)
    )
    sheet = (await _upload(setup, "omr1"))[0]
    assert sheet["match_status"] == "unmatched"

    resp = await setup.client.patch(
        f"/api/v1/omr/sheets/{sheet['id']}",
        json={"student_id": str(setup.students["99999"].id)},
    )

    assert resp.status_code == 200, resp.text
    data = resp.json()["data"]
    assert data["match_status"] == "manual"
    assert data["matched_manually"] is True
    assert data["student_id"] == str(setup.students["99999"].id)
    # A manual match must clear the "needs attention" reason it was flagged for.
    assert "Student match needs attention" not in (data["review_note"] or "")


async def test_assigning_an_already_claimed_student_conflicts(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    sheets = await _upload(setup, "omr1", "omr2")
    claimed_by = next(s for s in sheets if s["match_status"] == "matched")
    other = next(s for s in sheets if s["id"] != claimed_by["id"])

    resp = await setup.client.patch(
        f"/api/v1/omr/sheets/{other['id']}",
        json={"student_id": str(setup.students["11223"].id)},
    )

    assert resp.status_code == 409, resp.text
    assert "already matched to that student" in resp.json()["message"]


async def test_assigning_a_student_from_another_class_is_rejected(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    outsider = await _make_scannable_batch(
        db_session, role_clients, f"{unique_suffix}b", class_order=9, enrolled_rolls=("55555",)
    )
    sheet = (await _upload(setup, "omr2"))[0]

    resp = await setup.client.patch(
        f"/api/v1/omr/sheets/{sheet['id']}",
        json={"student_id": str(outsider.students["55555"].id)},
    )

    assert resp.status_code == 422, resp.text
    assert "not enrolled in this batch" in resp.json()["message"]


async def test_answer_override_changes_the_score(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    sheet = (await _upload(setup, "omr1"))[0]

    detail = (await setup.client.get(f"/api/v1/omr/sheets/{sheet['id']}")).json()["data"]
    # The key is a repeating Ka/Kha/Ga/Gha cycle, so question N's answer is known.
    options = ["Ka", "Kha", "Ga", "Gha"]
    wrong = [q for q, d in detail["score_details"].items() if d["status"] != "correct"]
    assert wrong, "fixture sheet should have at least one non-correct answer to fix"
    target = wrong[0]
    before = sheet["marks_obtained"]

    resp = await setup.client.patch(
        f"/api/v1/omr/sheets/{sheet['id']}",
        json={"answer_overrides": {target: options[(int(target) - 1) % 4]}},
    )

    assert resp.status_code == 200, resp.text
    data = resp.json()["data"]
    assert data["marks_obtained"] > before
    assert data["score_details"][target]["status"] == "correct"
    assert data["answers"][target]["confidence"] == "MANUAL"
    assert data["answers"][target]["overridden"] is True
    assert data["correct_count"] == sheet["correct_count"] + 1


async def test_override_option_and_question_are_validated(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    sheet = (await _upload(setup, "omr1"))[0]
    url = f"/api/v1/omr/sheets/{sheet['id']}"

    assert (await setup.client.patch(url, json={"answer_overrides": {"1": "Z"}})).status_code == 422
    assert (await setup.client.patch(url, json={"answer_overrides": {"999": "Ka"}})).status_code == 422
    assert (await setup.client.patch(url, json={})).status_code == 422


async def test_review_note_is_recorded_with_the_reviewer(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    from app.modules.omr.models import OmrSheet

    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    sheet = (await _upload(setup, "omr1"))[0]

    resp = await setup.client.patch(
        f"/api/v1/omr/sheets/{sheet['id']}",
        json={"review_note": "Smudged bubble on Q7, checked by hand"},
    )

    assert resp.status_code == 200, resp.text
    assert "Smudged bubble on Q7" in resp.json()["data"]["review_note"]

    stored = (
        await db_session.execute(select(OmrSheet).where(OmrSheet.id == uuid.UUID(sheet["id"])))
    ).scalars().one()
    await db_session.refresh(stored)
    assert stored.reviewed_by is not None
    assert stored.reviewed_at is not None


# --- Batch readiness ----------------------------------------------------------


async def test_batch_becomes_ready_once_everything_is_resolved(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("99999",)
    )
    sheet = (await _upload(setup, "omr1"))[0]

    batch = (await setup.client.get(f"/api/v1/omr/batches/{setup.batch_id}")).json()["data"]
    assert batch["status"] == "processing"

    # Resolve everything the scanner flagged: the identity, then any unclear answers.
    await setup.client.patch(
        f"/api/v1/omr/sheets/{sheet['id']}", json={"student_id": str(setup.students["99999"].id)}
    )
    detail = (await setup.client.get(f"/api/v1/omr/sheets/{sheet['id']}")).json()["data"]
    unclear = {
        question: "Ka"
        for question, entry in detail["answers"].items()
        if entry["status"] in ("multiple", "ambiguous")
    }
    if unclear:
        resp = await setup.client.patch(
            f"/api/v1/omr/sheets/{sheet['id']}", json={"answer_overrides": unclear}
        )
        assert resp.status_code == 200, resp.text

    final = (await setup.client.get(f"/api/v1/omr/sheets/{sheet['id']}")).json()["data"]
    assert final["status"] == "processed", final["review_note"]

    batch = (await setup.client.get(f"/api/v1/omr/batches/{setup.batch_id}")).json()["data"]
    assert batch["status"] == "ready"
    assert batch["matched_count"] == 1


async def test_deleting_the_last_sheet_returns_the_batch_to_draft(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    sheet = (await _upload(setup, "omr1"))[0]
    stored = set(cloudinary_stub.uploaded)

    resp = await setup.client.delete(f"/api/v1/omr/sheets/{sheet['id']}")

    assert resp.status_code == 200, resp.text
    assert set(cloudinary_stub.destroyed) == stored
    batch = (await setup.client.get(f"/api/v1/omr/batches/{setup.batch_id}")).json()["data"]
    assert batch["status"] == "draft"
    assert batch["sheet_count"] == 0
    assert (await setup.client.get(f"/api/v1/omr/sheets/{sheet['id']}")).status_code == 404


# --- Reprocess ----------------------------------------------------------------


async def test_reprocess_recovers_a_failed_sheet(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    """A sheet that failed on a bad upload is recoverable once the stored image
    is readable — without asking anyone to re-scan the paper."""
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    resp = await setup.client.post(
        f"/api/v1/omr/batches/{setup.batch_id}/sheets",
        files=[("images", ("truncated.jpg", b"\xff\xd8\xff\xe0 truncated", "image/jpeg"))],
    )
    sheet = resp.json()["data"]["sheets"][0]
    assert sheet["status"] == "failed"

    # Repair the stored asset, then reprocess.
    public_id = next(k for k in cloudinary_stub.uploaded if k.endswith(sheet["id"]))
    cloudinary_stub.uploaded[public_id] = (FIXTURES / "omr1.jpg").read_bytes()

    resp = await setup.client.post(f"/api/v1/omr/sheets/{sheet['id']}/reprocess")

    assert resp.status_code == 200, resp.text
    data = resp.json()["data"]
    assert data["status"] != "failed"
    assert data["error_message"] is None
    assert data["detected_roll"] == "011223"
    assert data["match_status"] == "matched"
    assert data["marks_obtained"] is not None


async def test_reprocess_preserves_a_manual_match_unless_reset(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("99999",)
    )
    sheet = (await _upload(setup, "omr1"))[0]
    student_id = str(setup.students["99999"].id)
    await setup.client.patch(f"/api/v1/omr/sheets/{sheet['id']}", json={"student_id": student_id})

    resp = await setup.client.post(f"/api/v1/omr/sheets/{sheet['id']}/reprocess")
    assert resp.status_code == 200, resp.text
    assert resp.json()["data"]["student_id"] == student_id
    assert resp.json()["data"]["match_status"] == "manual"

    resp = await setup.client.post(
        f"/api/v1/omr/sheets/{sheet['id']}/reprocess", params={"reset_match": "true"}
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()["data"]
    assert data["match_status"] == "unmatched"
    assert data["student_id"] is None
    assert data["matched_manually"] is False


# --- Access control -----------------------------------------------------------


async def test_review_routes_are_closed_to_other_teachers(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    sheet = (await _upload(setup, "omr1"))[0]

    resp = await role_clients["admin"].post(
        "/api/v1/teachers",
        json={
            "full_name": "OMR Review Outsider",
            "email": f"omr.review.{unique_suffix}@codexedumine.test",
            "phone": f"01955{unique_suffix[:6]}",
            "date_of_birth": "1990-01-01",
            "joining_date": str(date.today()),
        },
    )
    created = resp.json()["data"]

    other = AsyncClient(transport=ASGITransport(app=app), base_url="http://test")
    try:
        await other.post(
            "/api/v1/auth/login",
            json={"identifier": created["email"], "password": created["temporary_password"]},
        )
        url = f"/api/v1/omr/sheets/{sheet['id']}"
        assert (await other.patch(url, json={"review_note": "nope"})).status_code == 403
        assert (await other.post(f"{url}/reprocess")).status_code == 403
        assert (await other.delete(url)).status_code == 403
    finally:
        await other.aclose()


async def test_applied_batch_sheets_are_read_only(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    from app.common.enums import OmrBatchStatus
    from app.modules.omr.models import OmrBatch

    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    sheet = (await _upload(setup, "omr1"))[0]

    batch = (
        await db_session.execute(select(OmrBatch).where(OmrBatch.id == uuid.UUID(setup.batch_id)))
    ).scalars().one()
    await db_session.refresh(batch)
    batch.status = OmrBatchStatus.applied
    await db_session.commit()

    url = f"/api/v1/omr/sheets/{sheet['id']}"
    assert (await setup.client.patch(url, json={"review_note": "x"})).status_code == 409
    assert (await setup.client.post(f"{url}/reprocess")).status_code == 409
    assert (await setup.client.delete(url)).status_code == 409


# ==============================================================================
# Apply to the marks roster (Phase 7)
# ==============================================================================


async def _resolve_all_sheets(setup) -> None:
    """Clear every review flag a reviewer can legitimately clear, so the batch
    reaches "ready"."""
    resp = await setup.client.get(f"/api/v1/omr/batches/{setup.batch_id}/sheets")
    for sheet in resp.json()["data"]:
        detail = (await setup.client.get(f"/api/v1/omr/sheets/{sheet['id']}")).json()["data"]
        unclear = {
            question: "Ka"
            for question, entry in (detail["answers"] or {}).items()
            if entry["status"] in ("multiple", "ambiguous")
        }
        if unclear:
            resp = await setup.client.patch(
                f"/api/v1/omr/sheets/{sheet['id']}", json={"answer_overrides": unclear}
            )
            assert resp.status_code == 200, resp.text

    batch = (await setup.client.get(f"/api/v1/omr/batches/{setup.batch_id}")).json()["data"]
    assert batch["status"] == "ready", batch


async def _ready_batch(db_session, role_clients, suffix) -> SimpleNamespace:
    """A batch of class-10 fixture sheets, fully matched and resolved.

    Only omr2 and omr3 are used: omr1 is marked class 9, and a class mismatch is
    a verification flag the reviewer cannot clear (see the Phase 7 report note),
    so it could never reach "ready" in a class-10 batch.
    """
    setup = await _make_scannable_batch(
        db_session,
        role_clients,
        suffix,
        class_order=10,
        enrolled_rolls=("28637", "374789"),
        full_marks=40,
    )
    sheets = await _upload(setup, "omr2", "omr3")
    assert {s["match_status"] for s in sheets} == {"matched"}
    await _resolve_all_sheets(setup)
    return setup


async def test_apply_writes_marks_through_the_results_module(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    from app.modules.results.models import ExamResult

    setup = await _ready_batch(db_session, role_clients, unique_suffix)
    sheets = (await setup.client.get(f"/api/v1/omr/batches/{setup.batch_id}/sheets")).json()["data"]
    expected = {s["student_id"]: s["marks_obtained"] for s in sheets}

    resp = await setup.client.post(f"/api/v1/omr/batches/{setup.batch_id}/apply")

    assert resp.status_code == 200, resp.text
    data = resp.json()["data"]
    assert data["applied_count"] == 2
    assert data["unscanned"] == []
    assert data["skipped"] == []
    assert data["batch"]["status"] == "applied"
    assert data["batch"]["applied_at"] is not None

    rows = (
        await db_session.execute(
            select(ExamResult).where(ExamResult.exam_subject_id == setup.exam_subject.id)
        )
    ).scalars().all()
    assert len(rows) == 2
    for result in rows:
        await db_session.refresh(result)
        assert float(result.marks_obtained) == expected[str(result.student_id)]
        assert result.is_absent is False
        assert result.entered_by is not None
        # Grades are computed at publish time by results.service.publish_results,
        # not at mark entry — so they are still unset here, by design.
        assert result.grade is None

    # The sheets that produced those marks are now marked applied.
    sheets = (await setup.client.get(f"/api/v1/omr/batches/{setup.batch_id}/sheets")).json()["data"]
    assert {s["status"] for s in sheets} == {"applied"}


async def test_applied_marks_appear_on_the_teachers_own_roster(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _ready_batch(db_session, role_clients, unique_suffix)
    resp = await setup.client.post(f"/api/v1/omr/batches/{setup.batch_id}/apply")
    applied = {
        s["student_id"]: s["marks_obtained"]
        for s in (
            await setup.client.get(f"/api/v1/omr/batches/{setup.batch_id}/sheets")
        ).json()["data"]
    }

    resp = await setup.client.get(
        f"/api/v1/results/exam-subjects/{setup.exam_subject.id}/roster"
    )

    assert resp.status_code == 200, resp.text
    roster = resp.json()["data"]
    assert roster["full_marks"] == 40
    entered = {s["student_id"]: s["marks_obtained"] for s in roster["students"]}
    assert entered == applied


async def test_grades_are_computed_when_results_are_published(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    """End-to-end past apply: the OMR marks flow through the normal
    submit -> compile -> approve -> publish chain with no OMR-specific code."""
    from app.modules.results.models import ExamResult

    setup = await _ready_batch(db_session, role_clients, unique_suffix)
    await setup.client.post(f"/api/v1/omr/batches/{setup.batch_id}/apply")

    exam_id = str(
        (
            await db_session.execute(select(Exam).where(Exam.id == setup.exam_subject.exam_id))
        ).scalars().one().id
    )

    resp = await setup.client.post(
        f"/api/v1/results/exam-subjects/{setup.exam_subject.id}/submit"
    )
    assert resp.status_code == 200, resp.text
    resp = await role_clients["admin"].post(f"/api/v1/results/exams/{exam_id}/compile")
    assert resp.status_code == 200, resp.text
    resp = await role_clients["principal"].post(f"/api/v1/results/exams/{exam_id}/approve")
    assert resp.status_code == 200, resp.text
    resp = await role_clients["principal"].post(f"/api/v1/results/exams/{exam_id}/publish")
    assert resp.status_code == 200, resp.text

    rows = (
        await db_session.execute(
            select(ExamResult).where(ExamResult.exam_subject_id == setup.exam_subject.id)
        )
    ).scalars().all()
    assert len(rows) == 2
    for result in rows:
        await db_session.refresh(result)
        assert result.grade is not None, "publish_results should have graded the OMR marks"


async def test_double_apply_conflicts(role_clients, db_session, unique_suffix, cloudinary_stub):
    setup = await _ready_batch(db_session, role_clients, unique_suffix)

    assert (
        await setup.client.post(f"/api/v1/omr/batches/{setup.batch_id}/apply")
    ).status_code == 200

    resp = await setup.client.post(f"/api/v1/omr/batches/{setup.batch_id}/apply")
    assert resp.status_code == 409, resp.text
    assert "already been applied" in resp.json()["message"]


async def test_apply_requires_a_ready_batch(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("99999",)
    )

    # Nothing scanned yet.
    resp = await setup.client.post(f"/api/v1/omr/batches/{setup.batch_id}/apply")
    assert resp.status_code == 409, resp.text
    assert "'draft'" in resp.json()["message"]

    # Scanned but unresolved.
    await _upload(setup, "omr1")
    resp = await setup.client.post(f"/api/v1/omr/batches/{setup.batch_id}/apply")
    assert resp.status_code == 409, resp.text
    assert "'processing'" in resp.json()["message"]


async def test_apply_after_marks_are_submitted_surfaces_the_results_error(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    """The results module's own guard must be what stops this — and its own
    message must be what the user sees, not an OMR paraphrase."""
    setup = await _ready_batch(db_session, role_clients, unique_suffix)

    exam_subject = (
        await db_session.execute(
            select(ExamSubject).where(ExamSubject.id == setup.exam_subject.id)
        )
    ).scalars().one()
    await db_session.refresh(exam_subject)
    exam_subject.marks_submitted_at = datetime.now(timezone.utc)
    await db_session.commit()

    resp = await setup.client.post(f"/api/v1/omr/batches/{setup.batch_id}/apply")

    assert resp.status_code == 409, resp.text
    assert resp.json()["message"] == "Marks have already been submitted for this subject"

    # The failed apply must not have half-committed the batch.
    batch = (await setup.client.get(f"/api/v1/omr/batches/{setup.batch_id}")).json()["data"]
    assert batch["status"] == "ready"
    assert batch["applied_at"] is None


async def test_apply_after_the_deadline_surfaces_the_results_error(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _ready_batch(db_session, role_clients, unique_suffix)

    exam_subject = (
        await db_session.execute(
            select(ExamSubject).where(ExamSubject.id == setup.exam_subject.id)
        )
    ).scalars().one()
    await db_session.refresh(exam_subject)
    exam_subject.marks_deadline = datetime.now(timezone.utc) - timedelta(hours=1)
    await db_session.commit()

    resp = await setup.client.post(f"/api/v1/omr/batches/{setup.batch_id}/apply")

    assert resp.status_code == 422, resp.text
    assert resp.json()["message"] == "The marks submission deadline has passed"


async def test_students_without_a_sheet_are_reported_not_written(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    """A missing sheet means "not scanned", never "absent" — apply must not
    invent a zero for a student whose paper was not in the pile."""
    from app.modules.results.models import ExamResult

    setup = await _make_scannable_batch(
        db_session,
        role_clients,
        unique_suffix,
        class_order=10,
        enrolled_rolls=("28637", "374789", "55501"),
        full_marks=40,
    )
    await _upload(setup, "omr2", "omr3")
    await _resolve_all_sheets(setup)

    resp = await setup.client.post(f"/api/v1/omr/batches/{setup.batch_id}/apply")

    assert resp.status_code == 200, resp.text
    data = resp.json()["data"]
    assert data["applied_count"] == 2
    assert len(data["unscanned"]) == 1
    assert data["unscanned"][0]["roll_number"] == "55501"
    assert data["unscanned"][0]["student_id"] == str(setup.students["55501"].id)

    rows = (
        await db_session.execute(
            select(ExamResult).where(ExamResult.exam_subject_id == setup.exam_subject.id)
        )
    ).scalars().all()
    assert len(rows) == 2
    assert str(setup.students["55501"].id) not in {str(r.student_id) for r in rows}


async def test_eligibility_reports_an_applied_batch(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _ready_batch(db_session, role_clients, unique_suffix)

    resp = await setup.client.get(
        f"/api/v1/omr/exam-subjects/{setup.exam_subject.id}/eligibility"
    )
    assert resp.json()["data"]["has_applied_batch"] is False

    await setup.client.post(f"/api/v1/omr/batches/{setup.batch_id}/apply")

    resp = await setup.client.get(
        f"/api/v1/omr/exam-subjects/{setup.exam_subject.id}/eligibility"
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["data"]["has_applied_batch"] is True


async def test_apply_is_closed_to_other_teachers(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _ready_batch(db_session, role_clients, unique_suffix)

    resp = await role_clients["admin"].post(
        "/api/v1/teachers",
        json={
            "full_name": "OMR Apply Outsider",
            "email": f"omr.apply.{unique_suffix}@codexedumine.test",
            "phone": f"01966{unique_suffix[:6]}",
            "date_of_birth": "1990-01-01",
            "joining_date": str(date.today()),
        },
    )
    created = resp.json()["data"]

    other = AsyncClient(transport=ASGITransport(app=app), base_url="http://test")
    try:
        await other.post(
            "/api/v1/auth/login",
            json={"identifier": created["email"], "password": created["temporary_password"]},
        )
        resp = await other.post(f"/api/v1/omr/batches/{setup.batch_id}/apply")
        assert resp.status_code == 403, resp.text
    finally:
        await other.aclose()


# ==============================================================================
# Export (Phase 8)
# ==============================================================================


EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def test_csv_export_has_a_row_per_sheet(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    import csv as csv_module

    setup = await _make_scannable_batch(
        db_session,
        role_clients,
        unique_suffix,
        class_order=10,
        enrolled_rolls=("28637", "374789"),
        full_marks=40,
    )
    await _upload(setup, "omr2", "omr3")

    resp = await setup.client.get(
        f"/api/v1/omr/batches/{setup.batch_id}/export", params={"format": "csv"}
    )

    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith("text/csv")
    assert "attachment; filename=" in resp.headers["content-disposition"]
    assert resp.headers["content-disposition"].endswith('.csv"')

    rows = list(csv_module.reader(io.StringIO(resp.content.decode("utf-8-sig"))))
    assert rows[0][0] == "Student Name"
    assert len(rows) - 1 == 2  # header + one row per sheet

    names = {row[0] for row in rows[1:]}
    assert names == {"OMR Student 28637", "OMR Student 374789"}
    # Resolved student columns are populated, and the max-marks column is the
    # batch's MCQ ceiling rather than the answer key's own total.
    header = rows[0]
    assert rows[1][header.index("Max Marks")] == "40"
    assert rows[1][header.index("Class Roll")] in ("28637", "374789")


async def test_excel_export_has_the_three_expected_sheets(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    from openpyxl import load_workbook

    setup = await _make_scannable_batch(
        db_session,
        role_clients,
        unique_suffix,
        class_order=10,
        enrolled_rolls=("28637", "374789"),
        full_marks=40,
    )
    await _upload(setup, "omr2", "omr3")

    resp = await setup.client.get(
        f"/api/v1/omr/batches/{setup.batch_id}/export", params={"format": "excel"}
    )

    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(EXCEL_MEDIA_TYPE)
    assert resp.headers["content-disposition"].endswith('.xlsx"')
    assert len(resp.content) > 0

    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Summary", "Answer Details", "Statistics"]

    summary = wb["Summary"]
    # Row 1 is the title, row 2 the header, then one row per sheet.
    assert summary.max_row == 2 + 2
    assert [c.value for c in summary[2]][:3] == ["Student Name", "Admission No", "Class Roll"]

    answers = wb["Answer Details"]
    assert [c.value for c in answers[2]][1:] == [f"Q{i}" for i in range(1, 41)]

    stats = {row[0]: row[1] for row in wb["Statistics"].iter_rows(min_row=2, values_only=True) if row[0]}
    assert stats["Sheets in Batch"] == 2
    assert stats["Scored Sheets"] == 2


async def test_export_labels_unmatched_and_failed_sheets(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    """A sheet that could not be read or matched still appears — an export that
    silently dropped them would disagree with the batch it summarises."""
    import csv as csv_module

    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    files = _sheet_files("omr1", "omr2")
    files.append(("images", ("broken.jpg", b"not an image", "image/jpeg")))
    resp = await setup.client.post(f"/api/v1/omr/batches/{setup.batch_id}/sheets", files=files)
    assert resp.status_code == 201, resp.text

    resp = await setup.client.get(
        f"/api/v1/omr/batches/{setup.batch_id}/export", params={"format": "csv"}
    )

    rows = list(csv_module.DictReader(io.StringIO(resp.content.decode("utf-8-sig"))))
    assert len(rows) == 3

    by_file = {row["Source File"]: row for row in rows}
    assert by_file["broken.jpg"]["Status"] == "failed"
    assert by_file["broken.jpg"]["Student Name"] == "—"
    assert by_file["broken.jpg"]["Notes"]

    unmatched = by_file["omr2.jpg"]
    assert unmatched["Match"] == "unmatched"
    assert unmatched["Student Name"] == "—"
    # It was still read and scored; only the identity is missing.
    assert unmatched["Detected Roll"] == "028637"
    assert unmatched["Marks Obtained"] != "—"

    matched = by_file["omr1.jpg"]
    assert matched["Match"] == "matched"
    assert matched["Student Name"] == "OMR Student 11223"


async def test_export_filename_is_slugified(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(db_session, role_clients, unique_suffix)

    resp = await setup.client.get(
        f"/api/v1/omr/batches/{setup.batch_id}/export", params={"format": "csv"}
    )

    disposition = resp.headers["content-disposition"]
    filename = disposition.split('filename="', 1)[1].rstrip('"')
    assert re.fullmatch(r"[A-Za-z0-9_]+_omr_\d{4}-\d{2}-\d{2}\.csv", filename), filename


async def test_empty_batch_exports_headers_only(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    from openpyxl import load_workbook

    setup = await _make_scannable_batch(db_session, role_clients, unique_suffix)

    resp = await setup.client.get(
        f"/api/v1/omr/batches/{setup.batch_id}/export", params={"format": "excel"}
    )

    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Summary", "Answer Details", "Statistics"]
    assert wb["Summary"].max_row == 2
    assert wb["Answer Details"]["A1"].value == "No answer details available."


async def test_export_writes_nothing_to_disk(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    """Two uvicorn workers share no filesystem state — an exports/ directory
    would be written by one worker and missing from the other."""
    backend_dir = Path(__file__).parent.parent
    before = {p for p in backend_dir.rglob("*.xlsx")} | {p for p in backend_dir.rglob("*.csv")}

    setup = await _make_scannable_batch(
        db_session, role_clients, unique_suffix, class_order=9, enrolled_rolls=("11223",)
    )
    await _upload(setup, "omr1")

    for fmt in ("csv", "excel"):
        resp = await setup.client.get(
            f"/api/v1/omr/batches/{setup.batch_id}/export", params={"format": fmt}
        )
        assert resp.status_code == 200, resp.text

    after = {p for p in backend_dir.rglob("*.xlsx")} | {p for p in backend_dir.rglob("*.csv")}
    assert after == before


async def test_export_rejects_an_unknown_format(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(db_session, role_clients, unique_suffix)

    resp = await setup.client.get(
        f"/api/v1/omr/batches/{setup.batch_id}/export", params={"format": "pdf"}
    )

    assert resp.status_code == 422, resp.text


async def test_export_is_closed_to_other_teachers(
    role_clients, db_session, unique_suffix, cloudinary_stub
):
    setup = await _make_scannable_batch(db_session, role_clients, unique_suffix)

    resp = await role_clients["admin"].post(
        "/api/v1/teachers",
        json={
            "full_name": "OMR Export Outsider",
            "email": f"omr.export.{unique_suffix}@codexedumine.test",
            "phone": f"01977{unique_suffix[:6]}",
            "date_of_birth": "1990-01-01",
            "joining_date": str(date.today()),
        },
    )
    created = resp.json()["data"]

    other = AsyncClient(transport=ASGITransport(app=app), base_url="http://test")
    try:
        await other.post(
            "/api/v1/auth/login",
            json={"identifier": created["email"], "password": created["temporary_password"]},
        )
        resp = await other.get(f"/api/v1/omr/batches/{setup.batch_id}/export")
        assert resp.status_code == 403, resp.text
    finally:
        await other.aclose()
