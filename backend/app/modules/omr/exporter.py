"""
app/modules/omr/exporter.py
===========================
Exports a scanned OMR batch to Excel (3 sheets) and CSV.

Excel sheets:
  1. Summary        — one row per scanned sheet with student, metadata and score
  2. Answer Details — Q-by-Q answer status coloured green/red/yellow
  3. Statistics     — averages, grade distribution, pass rate, scan outcomes

Both writers return bytes rather than writing files. The deployment runs two
uvicorn workers, so an `exports/` directory on local disk would be written by
one worker and missing from the other; downloads stream straight from memory
instead.

Sheets that could not be read or could not be matched to a student are exported
too, labelled in the Status/Match columns. Dropping them would make the export
disagree with the batch it claims to summarise, and those are exactly the rows
someone reconciling a scan needs to see.
"""
from __future__ import annotations
import csv
import io
import logging
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Style constants ──────────────────────────────────────────────────────────
_HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
_CORR_FILL = PatternFill("solid", fgColor="C6EFCE")
_WRONG_FILL= PatternFill("solid", fgColor="FFC7CE")
_BLNK_FILL = PatternFill("solid", fgColor="FFEB9C")
_MULT_FILL = PatternFill("solid", fgColor="D9D9D9")
_ALT_FILL  = PatternFill("solid", fgColor="EBF3FB")

_HDR_FONT  = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
_BODY_FONT = Font(size=10, name="Calibri")
_TITLE_FONT= Font(bold=True, size=13, name="Calibri")

_THIN = Side(style="thin", color="BFBFBF")
_BORDER= Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER= Alignment(horizontal="center", vertical="center")
_LEFT  = Alignment(horizontal="left",   vertical="center")

UNRESOLVED = "—"

COLUMNS = [
    "Student Name", "Admission No", "Class Roll",
    "Status", "Match",
    "Detected Class", "Detected Roll", "Detected Subject Code", "Set Code",
    "Correct", "Wrong", "Blank", "Multiple",
    "Marks Obtained", "Max Marks", "Percentage (%)", "Grade",
    "Source File", "Notes",
]

_GRADE_COLUMN = COLUMNS.index("Grade") + 1


@dataclass(frozen=True)
class SheetExportRow:
    """One scanned sheet, flattened for export.

    Deliberately decoupled from the ORM so the writers below stay a pure
    formatting concern with no database knowledge.
    """

    student_name: str | None
    admission_number: str | None
    class_roll: str | None
    status: str
    match_status: str | None
    detected_class: int | None
    detected_roll: str | None
    detected_subject_code: str | None
    detected_set_code: str | None
    correct: int | None
    wrong: int | None
    blank: int | None
    multiple: int | None
    marks_obtained: float | None
    percentage: float | None
    grade: str | None
    source_file: str
    note: str | None
    score_details: dict[str, Any] | None

    @property
    def label(self) -> str:
        """How this row identifies itself when it has no resolved student."""
        if self.student_name:
            return self.student_name
        if self.detected_roll:
            return f"{UNRESOLVED} roll {self.detected_roll}"
        return f"{UNRESOLVED} {self.source_file}"

    def as_cells(self, max_marks: int) -> list[Any]:
        return [
            self.student_name or UNRESOLVED,
            self.admission_number or UNRESOLVED,
            self.class_roll or UNRESOLVED,
            self.status,
            self.match_status or UNRESOLVED,
            self.detected_class if self.detected_class is not None else UNRESOLVED,
            self.detected_roll or UNRESOLVED,
            self.detected_subject_code or UNRESOLVED,
            self.detected_set_code or UNRESOLVED,
            self.correct if self.correct is not None else UNRESOLVED,
            self.wrong if self.wrong is not None else UNRESOLVED,
            self.blank if self.blank is not None else UNRESOLVED,
            self.multiple if self.multiple is not None else UNRESOLVED,
            self.marks_obtained if self.marks_obtained is not None else UNRESOLVED,
            max_marks,
            self.percentage if self.percentage is not None else UNRESOLVED,
            self.grade or UNRESOLVED,
            self.source_file,
            (self.note or "").replace("\n", " | "),
        ]


class ResultExporter:
    """Renders a batch's sheets to CSV or Excel bytes."""

    def __init__(self, *, title: str, max_marks: int) -> None:
        self.title = title
        self.max_marks = max_marks

    # ------------------------------------------------------------------ #
    # CSV                                                                  #
    # ------------------------------------------------------------------ #

    def export_csv(self, rows: list[SheetExportRow]) -> bytes:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(COLUMNS)
        for row in rows:
            writer.writerow(row.as_cells(self.max_marks))
        # utf-8-sig so Excel opens Bengali student names correctly on Windows.
        return buffer.getvalue().encode("utf-8-sig")

    # ------------------------------------------------------------------ #
    # Excel                                                                #
    # ------------------------------------------------------------------ #

    def export_excel(self, rows: list[SheetExportRow]) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)  # type: ignore[arg-type]
        self._summary_sheet(wb, rows)
        self._answer_sheet(wb, rows)
        self._stats_sheet(wb, rows)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ── Sheet: Summary ───────────────────────────────────────────────────

    def _summary_sheet(self, wb: Workbook, rows: list[SheetExportRow]) -> None:
        ws = wb.create_sheet("Summary")
        n = len(COLUMNS)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
        tc = ws.cell(1, 1, self.title); tc.font = _TITLE_FONT; tc.alignment = _CENTER
        ws.row_dimensions[1].height = 26

        self._hdr_row(ws, 2, COLUMNS)

        for ri, row in enumerate(rows, 3):
            alt = _ALT_FILL if ri % 2 == 0 else None
            for ci, val in enumerate(row.as_cells(self.max_marks), 1):
                cell = ws.cell(ri, ci, val)
                cell.font = _BODY_FONT; cell.border = _BORDER; cell.alignment = _CENTER
                if alt:
                    cell.fill = alt
                if ci == _GRADE_COLUMN and isinstance(val, str):
                    cell.fill = (
                        _CORR_FILL if val in ("A+", "A") else
                        _WRONG_FILL if val == "F" else _ALT_FILL
                    )
            if row.status == "failed":
                for ci in range(1, n + 1):
                    ws.cell(ri, ci).fill = _WRONG_FILL

        self._auto_width(ws)

    # ── Sheet: Answer Details ────────────────────────────────────────────

    def _answer_sheet(self, wb: Workbook, rows: list[SheetExportRow]) -> None:
        ws = wb.create_sheet("Answer Details")

        q_numbers: list[str] = []
        for row in rows:
            if row.score_details:
                q_numbers = sorted(row.score_details.keys(), key=int)
                break
        if not q_numbers:
            ws["A1"] = "No answer details available."
            return

        nc = 1 + len(q_numbers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
        tc = ws.cell(1, 1, self.title); tc.font = _TITLE_FONT; tc.alignment = _CENTER
        ws.row_dimensions[1].height = 26

        self._hdr_row(ws, 2, ["Student"] + [f"Q{q}" for q in q_numbers])

        status_fill = {
            "correct": _CORR_FILL, "wrong": _WRONG_FILL,
            "blank": _BLNK_FILL,   "multiple": _MULT_FILL, "ambiguous": _MULT_FILL,
        }
        for ri, row in enumerate(rows, 3):
            details = row.score_details or {}
            cell = ws.cell(ri, 1, row.label)
            cell.font = _BODY_FONT; cell.border = _BORDER; cell.alignment = _LEFT

            for ci, q_str in enumerate(q_numbers, 2):
                entry = details.get(q_str, {})
                c = ws.cell(ri, ci, entry.get("student", UNRESOLVED))
                c.font = _BODY_FONT; c.border = _BORDER; c.alignment = _CENTER
                c.fill = status_fill.get(entry.get("status", ""), PatternFill())

        self._auto_width(ws, min_w=5, max_w=10)

    # ── Sheet: Statistics ────────────────────────────────────────────────

    def _stats_sheet(self, wb: Workbook, rows: list[SheetExportRow]) -> None:
        ws = wb.create_sheet("Statistics")

        scored = [r for r in rows if r.percentage is not None]
        pcts = [float(r.percentage) for r in scored]
        n = len(scored)
        avg = sum(pcts) / n if n else 0
        pass_rate = (sum(1 for p in pcts if p >= 40) / n * 100) if n else 0

        ws.merge_cells("A1:C1")
        tc = ws.cell(1, 1, f"{self.title} — Statistics")
        tc.font = _TITLE_FONT; tc.alignment = _CENTER
        ws.row_dimensions[1].height = 26

        stats = [
            ("Sheets in Batch",    len(rows)),
            ("Scored Sheets",      n),
            ("Unscored Sheets",    len(rows) - n),
            ("Average Score (%)",  round(avg, 2)),
            ("Highest Score (%)",  round(max(pcts), 2) if pcts else 0),
            ("Lowest Score (%)",   round(min(pcts), 2) if pcts else 0),
            ("Pass Rate (>=40%)",  f"{pass_rate:.1f}%"),
        ]
        for ri, (label, val) in enumerate(stats, 2):
            lc = ws.cell(ri, 1, label)
            lc.font = Font(bold=True, size=10, name="Calibri"); lc.border = _BORDER; lc.alignment = _LEFT
            vc = ws.cell(ri, 2, val)
            vc.font = _BODY_FONT; vc.border = _BORDER; vc.alignment = _CENTER

        start = len(stats) + 3
        ws.cell(start, 1, "Grade Distribution").font = Font(bold=True, size=10, name="Calibri")
        self._hdr_row(ws, start + 1, ["Grade", "Count", "Percentage"])
        grade_dist: dict[str, int] = {}
        for row in scored:
            grade = row.grade or UNRESOLVED
            grade_dist[grade] = grade_dist.get(grade, 0) + 1
        next_row = start + 2
        for grade, count in sorted(grade_dist.items()):
            ws.cell(next_row, 1, grade).border = _BORDER
            ws.cell(next_row, 2, count).border = _BORDER
            ws.cell(next_row, 3, f"{count / n * 100:.1f}%" if n else "0%").border = _BORDER
            next_row += 1

        # Scan outcomes: how every sheet ended up, so an incomplete export is
        # self-explanatory rather than looking like missing students.
        next_row += 1
        ws.cell(next_row, 1, "Scan Outcomes").font = Font(bold=True, size=10, name="Calibri")
        self._hdr_row(ws, next_row + 1, ["Match Result", "Count", "Percentage"])
        match_dist: dict[str, int] = {}
        for row in rows:
            key = row.match_status or row.status
            match_dist[key] = match_dist.get(key, 0) + 1
        total = len(rows)
        next_row += 2
        for label, count in sorted(match_dist.items()):
            ws.cell(next_row, 1, label).border = _BORDER
            ws.cell(next_row, 2, count).border = _BORDER
            ws.cell(next_row, 3, f"{count / total * 100:.1f}%" if total else "0%").border = _BORDER
            next_row += 1

        self._auto_width(ws)

    # ── Helpers ──────────────────────────────────────────────────────────

    @staticmethod
    def _hdr_row(ws: Any, row: int, headers: list[str]) -> None:
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row, ci, h)
            c.font = _HDR_FONT; c.fill = _HDR_FILL; c.border = _BORDER; c.alignment = _CENTER
        ws.row_dimensions[row].height = 18

    @staticmethod
    def _auto_width(ws: Any, min_w: int = 8, max_w: int = 28) -> None:
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0) for c in col_cells
            )
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(
                min_w, min(max_len + 3, max_w)
            )
